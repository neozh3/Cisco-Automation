from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'Switch Hostname.xlsx'

ws1 = wb.active
ws1.title = "Device Hotname Here"

switchDesc = [{'descrip': 'DESC1', 'port': 'Gi0/0', 'protocol': 'up', 'status': 'up'},
              {'descrip': 'DESC2', 'port': 'Gi0/1', 'protocol': 'up', 'status': 'up'},
              {'descrip': 'DESC3', 'port': 'Gi0/2', 'protocol': 'up', 'status': 'up'},
              {'descrip': 'DESCCCCCCCCC SUPER LONG AAAAAAAA',
               'port': 'Gi0/3',
               'protocol': 'up',
               'status': 'up'},
              {'descrip': '', 'port': 'Gi1/0', 'protocol': 'up', 'status': 'up'},
              {'descrip': '', 'port': 'Gi1/1', 'protocol': 'up', 'status': 'up'},
              {'descrip': '', 'port': 'Gi1/2', 'protocol': 'down', 'status': 'admin down'},
              {'descrip': '', 'port': 'Gi1/3', 'protocol': 'up', 'status': 'up'},
              {'descrip': 'DESCVLAN1',
               'port': 'Vl1',
               'protocol': 'down',
               'status': 'admin down'},
              {'descrip': '', 'port': 'Vl10', 'protocol': 'down', 'status': 'down'}]


ws1['B1'] = "hostname"
ws1.merge_cells('B1:M1')
keys = ['port', 'descrip', 'protocol', 'status']

colCounter = 0
# change 12to 14 for 12ports
for col in range(2, 12):
    # To count the 6 rows per interface
    rowCounter = 0
    # for the 6 rows per interface
    for row in range(2, 8):
        # print(colCounter)
        char = get_column_letter(col)
        if rowCounter == 0:
            ws1[char + str(row)] = switchDesc[colCounter]['port']
        elif rowCounter == 1:
            ws1[char + str(row)] = switchDesc[colCounter]['descrip']
        elif rowCounter == 5:
            ws1[char + str(row)] = switchDesc[colCounter]['status']
        else:
            ws1[char + str(row)] = None
        rowCounter += 1
    colCounter += 1


# for col in range(2, 14):
#    char = get_column_letter(col)
#    ws1.append(des)

# for row in range(2, 8):
#    ws1.append(des)

wb.save(filename=dest_filename)
