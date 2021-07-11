# This script reads all txt files in current directory and generates them into excel switchport assignments
# Currently only works on Cisco IOS / XE show interface description only
# txt file must be just the output of show int description, without the hostname prompt
# filename should be the hostname of the switch
# There can be multiple txt files in the folder

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import textfsm
import math
import os, time
import glob

templatePath = "C:\Program Files\ntc-templates-master\ntc_templates\templates"

def convertTxtfsmOp(header, result):
    '''convert textfsm output to netmiko textfsm output(JSON)'''
    output = []
    for n in range(len(result)):
        res = {header[i].lower(): result[n][i] for i in range(len(header))}
        output.append(res)

    return output

# function to calculate the range of cell values for the rows
def calcRanges(twelvePortRowsNeeded):
    startValue = 2
    rowPerIface = 6
    rangeList = []

    for n in range(twelvePortRowsNeeded):
        endValue = startValue + rowPerIface
        r = range(startValue, endValue)
        rangeList.append(r)
        startValue = endValue
        endValue = endValue + rowPerIface

    return rangeList

# ----------- Styles-------------------
ft = Font(name="Arial", size=8)

align = Alignment(horizontal="center", vertical="center")

topBdr = Border(top=Side(border_style='thick'), left=Side(
    border_style='thick'), right=Side(border_style='thick'))
sideBdr = Border(left=Side(border_style='thick'), right=Side(border_style='thick'))
btmBdr = Border(bottom=Side(border_style='thick'), left=Side(
    border_style='thick'), right=Side(border_style='thick'))

redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
greenFill = PatternFill(start_color="0099CC00", end_color = "0099CC00", fill_type="solid")
amberFill = PatternFill(start_color="00FF9900", end_color = "00FF9900", fill_type="solid")
greyFill = PatternFill(start_color="00C0C0C0", end_color = "00C0C0C0", fill_type="solid")
# ----------- End Styles-------------------


# read all txt files in current folder
for txtFilename in glob.glob('*.txt'):

    hostname = txtFilename.strip('.txt')
    wb = Workbook()
    dest_filename = f'{hostname}.xlsx'
    ws1 = wb.active
    ws1.title = f"{hostname}"

    # parse interface description from text file into json/ textfsm format
    with open(rf"C:\Program Files\ntc-templates-master\ntc_templates\templates\cisco_ios_show_interfaces_description.textfsm") as template, open(txtFilename, "r") as output:

        table = textfsm.TextFSM(template)

        header = table.header

        result = table.ParseText(output.read())
        switchDesc = convertTxtfsmOp(header, result)
        lastModified = os.path.getmtime(txtFilename)
        lastModified = time.strftime('%d-%b-%Y', time.localtime(lastModified))

    # add code here to filter switchDesc eg. exclude vlan & portchannel interfaces
    tmp = []

    for iface in switchDesc:
        if iface['port'][0:2] == 'V1':
            tmp.append(iface)
        elif iface['port'][0:2] == 'Po':
            tmp.append(iface)

    newDesc = []

    for ele in switchDesc:
        if ele not in tmp:
            newDesc.append(ele)

    switchDesc = newDesc

    # find the total number of 12 port rows needed, rounded up
    twelvePortRowsNeeded = math.ceil(len(switchDesc)/12)

    # for each row, script will loop 12x, so need to add in filler values to allow
    # total number of interfaces to be a multiple of 12. If not there will be error
    fillerVal = {'port': "", "status": "", "protocol": "", "descrip": ""}
    numOfFillerReq = (twelvePortRowsNeeded * 12) - len(switchDesc)
    for z in range(numOfFillerReq):
        switchDesc.append(fillerVal)

    rowRanges = calcRanges(twelvePortRowsNeeded)


    # cell B1 merge till M1 as the hostname of the switch
    ws1["B1"] = f"{hostname}"

    ws1.merge_cells('B1:M1')

    # -----------to populate the excel with the port descriptions etc-------------

    # port number of the switch, starting from 0
    portCounter = 0
    for p in range(twelvePortRowsNeeded):

        # for column b to m
        for col in range(2, 14):
            char = get_column_letter(col)

            # portSectionCounter is the "port" in the excel, from top down 0-6
            portSectionCounter = 0

            for portRowExcel in rowRanges[p]:
                coor = char + str(portRowExcel)
                ws1[coor].font = ft
                ws1[coor].alignment = align
                #To get colour to fill based on interface status
                if switchDesc[portCounter]['status'] == 'up':
                    toFill = greenFill
                elif switchDesc[portCounter]['status'] == 'down':
                    toFill = amberFill
                else:
                    toFill = redFill

                if portSectionCounter == 0:
                    ws1[coor].border = topBdr
                    ws1[coor] = switchDesc[portCounter]['port']
                    ws1[coor].fill = greyFill

                if portSectionCounter == 1:
                    ws1[coor].border = sideBdr
                    ws1[coor] = switchDesc[portCounter]['descrip']
                    ws1[coor].fill = toFill

                if portSectionCounter == 2:
                    ws1[coor].border = sideBdr
                    ws1[coor].fill = toFill

                if portSectionCounter == 3:
                    ws1[coor].border = sideBdr
                    ws1[coor].fill = toFill

                if portSectionCounter == 4:
                    ws1[coor].border = sideBdr
                    ws1[coor] = switchDesc[portCounter]['status']
                    ws1[coor].fill = toFill

                if portSectionCounter == 5:
                    ws1[coor].border = btmBdr
                    ws1[coor].fill = toFill

                portSectionCounter += 1

            portCounter += 1

    #------Legend portion ------------------------
    dateModCellRow = rowRanges[-1][-1] + 2
    ws1[f'A{dateModCellRow}'] = f'Date Updated: {lastModified}'

    colourLegendRow1 = dateModCellRow + 1
    colourLegendRow2 = dateModCellRow + 2
    colourLegendRow3 = dateModCellRow + 3

    ws1[f'A{colourLegendRow1}'].fill = greenFill
    ws1[f'A{colourLegendRow1}'] = "Active Connection"

    ws1[f'A{colourLegendRow2}'].fill = amberFill
    ws1[f'A{colourLegendRow2}'] = "Interface Down"

    ws1[f'A{colourLegendRow3}'].fill = redFill
    ws1[f'A{colourLegendRow3}'] = "Interface Shutdown"
    #------End Legend portion ------------------------

    #------all descriptions portion ------------------------
    allSvrListCol = 'O'
    allSvrListRow = 2

    ws1[f'{allSvrListCol}1'] = "All Description:"
    for i in range(0,len(switchDesc)):
        r = str(allSvrListRow)
        ws1[f'{allSvrListCol}{r}'] = switchDesc[i]['descrip']
        allSvrListRow +=  1
    #------End all descriptions portion ------------------------

    wb.save(filename=dest_filename)
